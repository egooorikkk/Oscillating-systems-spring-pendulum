import tkinter as tk
from tkinter import ttk, messagebox
import numpy as np 
from openpyxl import Workbook, load_workbook  # Библиотека для работы с Excel-файлами (чтение, запись).
from matplotlib.figure import Figure  
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import simpledialog, filedialog  # Модули для диалоговых окон: simpledialog для ввода данных, filedialog для выбора файлов.
import os  # Для работы с файлами
import math 
from Email_sender import EmailSender 
from Animation import Animation 
from AuthorInfoWindow import AuthorInfoWindow  
from ProgramInfoWindow import ProgramInfoWindow  
from TheoryWindow import TheoryWindow 
from UserGuideWindow import UserGuideWindow 


class MainWindow:
    def __init__(self, app):
        """Инициализация основного окна."""
        self.app = app
        self.main_window = tk.Toplevel()
        self.main_window.title("Основное окно")
        self.main_window.geometry("1900x950")
        self.create_menu()  # Создание меню
        self.create_widgets()  # Создание интерфейса
    
        self.previous_data = {} # Хранение данных для восстановления

    def create_widgets(self):
        """Создание виджетов для ввода, управления и отображения данных."""
        # Фрейм для текстовых полей
        text_frame = tk.Frame(self.main_window)
        text_frame.pack(pady=10)

        ttk.Label(text_frame, text="Масса груза (кг):", font=("Arial", 14)).grid(row=0, column=0, padx=20)
        ttk.Label(text_frame, text="Жесткость пружины (Н/м):", font=("Arial", 14)).grid(row=0, column=1, padx=20)

        # Поля ввода
        input_frame = tk.Frame(self.main_window)
        input_frame.pack(pady=10)

        self.mass_entry = ttk.Entry(input_frame, font=("Arial", 14), width=15)
        self.mass_entry.grid(row=1, column=0, padx=20)

        self.spring_entry = ttk.Entry(input_frame, font=("Arial", 14), width=15)
        self.spring_entry.grid(row=1, column=1, padx=20)

        # Кнопки для расчёта и очистки полей
        calc_clear_frame = tk.Frame(self.main_window)
        calc_clear_frame.pack(pady=10)

        calculate_button = tk.Button(calc_clear_frame, text="Запустить расчеты", command=self.calculate_and_plot, font=("Arial", 14))
        calculate_button.grid(row=2, column=0, padx=10)

        clear_button = tk.Button(calc_clear_frame, text="Очистить поля", command=self.clear_entries, font=("Arial", 14))
        clear_button.grid(row=2, column=1, padx=10)

        # Фрейм для графиков
        self.graph_frame = tk.Frame(self.main_window)
        self.graph_frame.pack(pady=20, expand=True, fill="both")

        # Метка для вывода результатов
        self.result_label = ttk.Label(self.graph_frame, text="", font=("Arial", 14))
        self.result_label.pack(pady=10)

        # Фрейм для дополнительных кнопок
        button_frame = tk.Frame(self.main_window)
        button_frame.pack(side="bottom", pady=10)

        author_button = tk.Button(button_frame, text="Об авторе", command=self.show_author_info, font=("Arial", 14))
        author_button.pack(side="left", padx=(0, 10))

        program_button = tk.Button(button_frame, text="О программе", command=self.show_program_info, font=("Arial", 14))
        program_button.pack(side="left", padx=(0, 10))

        save_button = tk.Button(button_frame, text="Сохранить график", command=self.save_graph, font=("Arial", 14))
        save_button.pack(side="left", padx=(0, 10))

        animation_button = tk.Button(self.main_window, text="Мультипликация движения", command=self.open_animation_window, font=("Arial", 14))
        animation_button.pack(pady=20)

        export_button = tk.Button(button_frame, text="Экспорт данных", command=self.export_to_excel, font=("Arial", 14))
        export_button.pack(side="left", padx=(0, 10))

        exit_button_main = tk.Button(button_frame, text="Выход", command=self.main_window.quit, font=("Arial", 14))
        exit_button_main.pack(side="left", padx=(0, 10))

    def create_menu(self):
        """Создание главного меню приложения."""
        # Создаем меню
        menu_bar = tk.Menu(self.main_window)

        # Меню "Файл"
        file_menu = tk.Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="Открыть папку с изображением графиков", command=self.open_file)
        file_menu.add_command(label="Отправка графика по электронной почте", command=self.gmail)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.main_window.quit)
        menu_bar.add_cascade(label="Файл", menu=file_menu)
        self.main_window.config(menu=menu_bar)

        # Меню "Справка"
        help_menu = tk.Menu(menu_bar, tearoff=0)
        help_menu.add_command(label="Руководство пользователя", command=self.show_user_guide)
        help_menu.add_command(label="Теория", command=self.show_Theory)
        help_menu.add_separator()
        help_menu.add_command(label="Об авторе", command=self.show_author_info)
        help_menu.add_command(label="О программе", command=self.show_program_info)
        menu_bar.add_cascade(label="Справка", menu=help_menu)

        # Меню "Восстановление"
        restore_menu = tk.Menu(menu_bar, tearoff=0)
        restore_menu.add_command(label="Восстановить предыдущее состояние", command=self.restore_previous_state)
        menu_bar.add_cascade(label="Восстановление", menu=restore_menu)


    def restore_previous_state(self):
        if self.previous_data:
            # Восстанавливаем данные
            self.mass_entry.delete(0, tk.END)
            self.mass_entry.insert(0, self.previous_data['mass'])
            self.spring_entry.delete(0, tk.END)
            self.spring_entry.insert(0, self.previous_data['spring'])

            # Восстанавление периода
            if 'periods' in self.previous_data:
                self.periods = self.previous_data['periods']
                periods_text = ", ".join(f"{p:.2f}" for p in self.periods)

                # Удаление старой метки с периодами, если она существует
                if hasattr(self, 'result_label'):
                    self.result_label.destroy()

                self.result_label = ttk.Label(self.graph_frame, text=f"Периоды колебаний: {periods_text} секунд", font=("Arial", 14))
                self.result_label.pack(pady=10)

            # Восстанавливаем график
            if 'figure' in self.previous_data and self.previous_data['figure']:
                for widget in self.graph_frame.winfo_children():
                    widget.destroy()

                canvas = FigureCanvasTkAgg(self.previous_data['figure'], master=self.graph_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            messagebox.showinfo("Восстановление", "Предыдущее состояние успешно восстановлено.")
        else:
            messagebox.showwarning("Ошибка восстановления", "Нет сохранённых данных для восстановления.")

    def show_user_guide(self):
        UserGuideWindow(self.main_window)


    def gmail(self):
        # Открытие диалога для выбора файла
        file_path = filedialog.askopenfilename(
            title="Выберите файл с графиком", 
            filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.gif")]
        )
        
        if not file_path:
            return
        
        # Запрос на ввод данных для письма
        to_email = simpledialog.askstring("Получатель", "Введите адрес электронной почты получателя:")
        subject = simpledialog.askstring("Тема", "Введите тему письма:")
        body = simpledialog.askstring("Текст письма", "Введите текст письма:")
        
        # Отправка письма
        email_sender = EmailSender()
        email_sender.send_email(file_path, to_email, subject, body)


    def open_theory_window(self):
         messagebox.showinfo("Сброс", "Настройки будут сброшены к стандартным.")

    def open_file(self):
        folder_path = r"E:\Программирование\new\graphics"
    
        # Проверка существования папки
        if not os.path.exists(folder_path):
            messagebox.showerror("Ошибка", f"Папка '{folder_path}' не найдена!")
            return
        
        # Открытие проводника Windows
        os.startfile(folder_path)

    def save_file(self):
        messagebox.showinfo("Сохранить файл", "Здесь можно добавить логику для сохранения файла.")

    def open_animation_window(self):
        animation = Animation()
        animation.run()

    def export_data(self):
        messagebox.showinfo("Сохранить файл", "Здесь можно добавить логику для сохранения файла.")

    def clear_entries(self):
        # Очищаем только поля ввода
        self.mass_entry.delete(0, tk.END)
        self.spring_entry.delete(0, tk.END)

    def calculate_and_plot(self):
        """Рассчёт данных и построение нового графика."""
        mass_str = self.mass_entry.get()
        spring_str = self.spring_entry.get()

        try:
            mass = float(mass_str)
            spring_constant = float(spring_str)

            if mass <= 0 or spring_constant <= 0:
                raise ValueError

            # Сохраняем текущее состояние перед расчётом нового графика
            if hasattr(self, 'app') and hasattr(self.app, 'current_fig') and self.app.current_fig is not None:
                self.previous_data = {
                    'mass': self.mass_entry.get(),
                    'spring': self.spring_entry.get(),
                    'periods': getattr(self, 'periods', []),
                    'figure': self.app.current_fig
                }

            # Рассчитываем данные
            period = 2 * np.pi * np.sqrt(mass / spring_constant)
            messagebox.showinfo("Период колебаний", f"Период колебаний: {period:.2f} секунд")

            # Инициализируем или обновляем список периодов
            if not hasattr(self, 'periods'):
                self.periods = []
            self.periods.append(period)

            # Форматируем текст периодов
            periods_text = ", ".join(f"{p:.2f}" for p in self.periods)

            # Удаляем старую метку с периодами, если она существует
            if hasattr(self, 'result_label'):
                self.result_label.destroy()

            self.result_label = ttk.Label(self.graph_frame, text=f"Периоды колебаний: {periods_text} секунд", font=("Arial", 14))
            self.result_label.pack(pady=10)

            # Генерируем данные для графиков
            t = np.linspace(0, 10, 100)
            displacement = np.sin(np.sqrt(spring_constant / mass) * t)
            velocity = np.sqrt(spring_constant / mass) * np.cos(np.sqrt(spring_constant / mass) * t)
            acceleration = -spring_constant / mass * displacement

            # Удаляем старые графики
            for widget in self.graph_frame.winfo_children():
                widget.destroy()

            # Создаём новый график
            self.app.current_fig = Figure(figsize=(10, 6))
            ax1 = self.app.current_fig.add_subplot(311)
            ax1.plot(t, displacement, label='Смещение', color='blue')
            ax1.set_title('Смещение от времени')
            ax1.set_ylabel('Смещение (м)')
            ax1.legend()

            ax2 = self.app.current_fig.add_subplot(312)
            ax2.plot(t, velocity, label='Скорость', color='green')
            ax2.set_title('Скорость от времени')
            ax2.set_ylabel('Скорость (м/с)')
            ax2.legend()

            ax3 = self.app.current_fig.add_subplot(313)
            ax3.plot(t, acceleration, label='Ускорение', color='red')
            ax3.set_title('Ускорение от времени')
            ax3.set_ylabel('Ускорение (м/с²)')
            ax3.set_xlabel('Время (с)')
            ax3.legend()

            # Отображаем график
            canvas = FigureCanvasTkAgg(self.app.current_fig, master=self.graph_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        except ValueError:
            messagebox.showerror("Ошибка", "Введите положительные числовые значения для массы и жёсткости.")
    

    def save_graph(self):
        """Сохранение графика в файл."""
        # Проверка на наличие атрибута current_fig у self.app
        if hasattr(self.app, 'current_fig') and self.app.current_fig is not None:
            file_path = filedialog.asksaveasfilename(defaultextension=".png", 
                                                    filetypes=[("PNG files", "*.png"), 
                                                            ("All files", "*.*")])
            if file_path:
                self.app.current_fig.savefig(file_path)
                messagebox.showinfo("Сохранение", "График успешно сохранён!")
        else:
            messagebox.showerror("Ошибка", "График отсутствует. Пожалуйста, создайте график перед сохранением.")


    
    def export_to_excel(self):
        """Экспорт результатов расчёта в файл Excel."""

        try:
            # Получаем введённые значения
            mass = float(self.mass_entry.get())  # Масса
            spring_constant = float(self.spring_entry.get())  # Жёсткость пружины

            if mass <= 0 or spring_constant <= 0:
                raise ValueError("Параметры должны быть больше нуля.")

            # Вычисляем период
            period = 2 * math.pi * math.sqrt(mass / spring_constant)

            # Упрощённый расчёт амплитуды
            amplitude = 0.5 * math.sqrt(mass / spring_constant)  # Условное приближение амплитуды

            # Дополнительные вычисления
            frequency = 1 / period  # Частота
            max_acceleration = amplitude * (2 * math.pi / period) ** 2  # Максимальное ускорение
            energy = 0.5 * spring_constant * amplitude ** 2  # Механическая энергия

            # Выбор файла для сохранения или открытия
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Сохранить или выбрать файл"
            )

            if not file_path:
                return  # Если пользователь отменил выбор

            # Если файл существует, открываем его, иначе создаём новый
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Результаты"
                # Добавляем заголовки для нового файла
                headers = ["Масса (кг)", "Жёсткость (Н/м)", "Период (с)", "Частота (Гц)", "Амплитуда (м)", "Макс. ускорение (м/с²)", "Энергия (Дж)"]
                ws.append(headers)

            # Добавляем новую строку данных
            ws.append([mass, spring_constant, round(period, 2), round(frequency, 2), round(amplitude, 2), round(max_acceleration, 2), round(energy, 2)])

            # Сохраняем изменения
            wb.save(file_path)
            messagebox.showinfo("Экспорт", f"Данные успешно сохранены!")

        except ValueError as e:
            messagebox.showerror("Ошибка", f"Некорректные данные: {e}")
        except Exception as e:
            messagebox.showerror("Ошибка экспорта", f"Произошла ошибка: {e}")

    def show_author_info(self):
        """Об авторе"""
        self.author_info_window = AuthorInfoWindow(self.main_window)

    def show_program_info(self):
        """О программе"""
        self.program_info_window = ProgramInfoWindow(self.main_window) 

    def show_Theory(self):
        """Теория"""
        self.theory_window = TheoryWindow(self.main_window)  
